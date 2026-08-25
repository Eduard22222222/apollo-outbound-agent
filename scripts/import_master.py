"""Import the proprietary target-company database into the LOCAL master.

This file never goes to Apollo (HARD RULE H1). Turnover, profit, CAEN, source and notes
are stored here and stripped from anything pushed later (docs/02 sec.3).

    python scripts/import_master.py --file data/raw/companies.csv           # dry run
    python scripts/import_master.py --file data/raw/companies.csv --execute

Column names are matched loosely and case-insensitively, EN and RO:
  name / denumire / company        domain / website / site
  cui / cif / vat                  caen / cod caen
  employees / angajati             turnover / cifra afaceri / ca
  profit / profit net              city / oras / localitate
  email  first_name/prenume  last_name/nume  title/functie
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from common import (
    REPO_ROOT,
    db_connect,
    die,
    load_config,
    log_audit,
    norm_cui,
    norm_domain,
    norm_email,
    norm_name,
    out,
    require_execute,
    split_name,
    table,
    utcnow,
)

ALIASES = {
    "name": ["name", "company", "company_name", "denumire", "denumire firma", "firma",
             "nume firma", "organization", "organization_name"],
    "domain": ["domain", "website", "web", "site", "url", "domeniu", "pagina web"],
    "cui": ["cui", "cif", "cod fiscal", "vat", "vat_id", "tax_id"],
    "caen": ["caen", "cod caen", "cod_caen", "activitate", "industry_code"],
    "employees": ["employees", "employee_count", "angajati", "nr angajati", "numar angajati",
                  "headcount"],
    "turnover": ["turnover", "revenue", "cifra de afaceri", "cifra afaceri", "ca", "venituri"],
    "profit": ["profit", "profit net", "net profit", "net_income", "profit brut"],
    "fiscal_year": ["year", "an", "fiscal_year", "an fiscal", "exercitiu"],
    "city": ["city", "oras", "localitate", "municipiu"],
    "country": ["country", "tara"],
    "linkedin_url": ["linkedin", "linkedin_url", "linkedin url"],
    "segment": ["segment", "industry", "industrie", "vertical", "sector"],
    "notes": ["notes", "note", "observatii", "comentarii"],
    "email": ["email", "e-mail", "mail", "contact_email"],
    "first_name": ["first_name", "prenume", "firstname"],
    "last_name": ["last_name", "nume", "lastname", "surname"],
    "full_name": ["full_name", "contact", "persoana", "nume complet", "contact_name"],
    "title": ["title", "functie", "job_title", "pozitie", "role"],
    "phone": ["phone", "telefon", "mobil", "mobile"],
}


def build_map(headers: list[str]) -> dict[str, str]:
    lowered = {h.strip().lower(): h for h in headers}
    mapping: dict[str, str] = {}
    for field, names in ALIASES.items():
        for candidate in names:
            if candidate in lowered:
                mapping[field] = lowered[candidate]
                break
    return mapping


def read_rows(path: Path) -> tuple[list[dict], list[str]]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook  # optional dependency
        except ImportError:
            die("XLSX needs openpyxl: pip install openpyxl - or export the sheet to CSV")
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        headers = [str(c).strip() if c is not None else "" for c in next(it)]
        rows = [
            {headers[i]: ("" if v is None else str(v)) for i, v in enumerate(r) if i < len(headers)}
            for r in it
        ]
        return rows, headers
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        return list(reader), list(reader.fieldnames or [])


def num(value) -> float | None:
    if value in (None, ""):
        return None
    txt = str(value).replace(" ", "").replace(" ", "").replace(",", ".")
    txt = "".join(ch for ch in txt if ch.isdigit() or ch in ".-")
    try:
        return float(txt)
    except ValueError:
        return None


def main() -> int:
    parser = argparse.ArgumentParser(description="Import companies/contacts into the local master")
    parser.add_argument("--file", required=True)
    parser.add_argument("--source", default="", help="where this list came from (kept LOCAL)")
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--dry-run", action="store_true",
                        help="(default) preview only; kept so the flag is explicit")
    parser.add_argument("--limit", type=int, default=0, help="only the first N rows")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.is_absolute():
        path = REPO_ROOT / path
    if not path.exists():
        die(f"no such file: {path}")

    config = load_config()
    conn = db_connect(config)

    rows, headers = read_rows(path)
    if args.limit:
        rows = rows[: args.limit]
    if not rows:
        die("file has no data rows")

    mapping = build_map(headers)
    if "name" not in mapping and "domain" not in mapping:
        die(f"could not find a company name or domain column in: {headers}")

    out(f"{path.name}: {len(rows)} rows")
    out("column mapping:")
    for field, col in sorted(mapping.items()):
        out(f"  {field:<12} <- {col}")
    unmapped = [h for h in headers if h not in mapping.values() and h]
    if unmapped:
        out(f"  ignored: {', '.join(unmapped[:12])}{' ...' if len(unmapped) > 12 else ''}")

    stats = {"companies_new": 0, "companies_dup": 0, "contacts_new": 0,
             "contacts_dup": 0, "skipped": 0}
    preview: list[dict] = []
    # In-run keys, so a dry run reports the same dedupe result an --execute run would.
    seen_domains: set[str] = set()
    seen_cuis: set[str] = set()
    seen_names: set[str] = set()
    seen_emails: set[str] = set()

    for row in rows:
        def get(field: str, row=row) -> str:   # row bound per iteration, not by closure
            col = mapping.get(field)
            return (row.get(col) or "").strip() if col else ""

        name = get("name")
        domain = norm_domain(get("domain"))
        cui = norm_cui(get("cui"))
        if not name and not domain:
            stats["skipped"] += 1
            continue
        name = name or domain
        nname = norm_name(name)

        dup_in_run = bool(
            (domain and domain in seen_domains)
            or (cui and cui in seen_cuis)
            or (nname and nname in seen_names)
        )
        existing = None
        if domain:
            existing = conn.execute(
                "SELECT id FROM companies WHERE domain = ?", (domain,)
            ).fetchone()
        if existing is None and cui:
            existing = conn.execute("SELECT id FROM companies WHERE cui = ?", (cui,)).fetchone()
        if existing is None and nname:
            existing = conn.execute(
                "SELECT id FROM companies WHERE name_norm = ?", (nname,)
            ).fetchone()

        if domain:
            seen_domains.add(domain)
        if cui:
            seen_cuis.add(cui)
        if nname:
            seen_names.add(nname)

        if existing or dup_in_run:
            stats["companies_dup"] += 1
            company_id = existing["id"] if existing else None
        else:
            stats["companies_new"] += 1
            company_id = None
            if args.execute:
                cur = conn.execute(
                    "INSERT INTO companies (name, name_norm, domain, cui, caen, country, city,"
                    " employees, turnover, profit, fiscal_year, website, linkedin_url, source,"
                    " notes, segment, first_seen, updated_at)"
                    " VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    (
                        name, nname, domain, cui, get("caen") or None,
                        get("country") or "RO", get("city") or None,
                        int(num(get("employees")) or 0) or None,
                        num(get("turnover")), num(get("profit")),
                        int(num(get("fiscal_year")) or 0) or None,
                        get("domain") or None, get("linkedin_url") or None,
                        args.source or path.name, get("notes") or None,
                        get("segment") or None, utcnow(), utcnow(),
                    ),
                )
                company_id = cur.lastrowid
            if len(preview) < 5:
                preview.append({"name": name, "domain": domain or "-", "cui": cui or "-",
                                "caen": get("caen") or "-", "city": get("city") or "-"})

        email = norm_email(get("email"))
        first, last = get("first_name"), get("last_name")
        if not first and not last:
            first, last = split_name(get("full_name"))
        if not (email or first or last):
            continue

        dup = None
        if email:
            if email in seen_emails:
                dup = True
            else:
                dup = conn.execute(
                    "SELECT id FROM contacts WHERE email = ?", (email,)
                ).fetchone()
            seen_emails.add(email)
        if dup:
            stats["contacts_dup"] += 1
            continue
        stats["contacts_new"] += 1
        if args.execute and company_id:
            conn.execute(
                "INSERT INTO contacts (company_id, first_name, last_name, full_name, title,"
                " email, phone, linkedin_url, source, first_seen, updated_at)"
                " VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                (
                    company_id, first or None, last or None,
                    (f"{first} {last}".strip() or None), get("title") or None,
                    email, get("phone") or None, get("linkedin_url") or None,
                    args.source or path.name, utcnow(), utcnow(),
                ),
            )

    out("")
    if preview:
        out("sample of new companies:")
        out(table(preview, ["name", "domain", "cui", "caen", "city"]))
        out("")
    out(f"companies: {stats['companies_new']} new, {stats['companies_dup']} already known")
    out(f"contacts : {stats['contacts_new']} new, {stats['contacts_dup']} already known")
    out(f"skipped  : {stats['skipped']} rows with neither name nor domain")

    if not require_execute(args, "write these rows into the local master"):
        return 0

    conn.commit()
    log_audit(conn, "operator", "import_master", path.name, dry_run=False, detail=str(stats))

    report = REPO_ROOT / "docs" / "reports" / f"import_{utcnow()[:10]}.md"
    report.parent.mkdir(parents=True, exist_ok=True)
    with report.open("a", encoding="utf-8") as fh:
        fh.write(f"\n## {path.name} - {utcnow()}\n\n")
        for key, value in stats.items():
            fh.write(f"- {key}: {value}\n")
        fh.write(f"- source label: {args.source or path.name}\n")
    out(f"\nwritten. report appended to {report.relative_to(REPO_ROOT)}")
    out("Reminder: this data stays local. Apollo only ever receives the current batch.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
